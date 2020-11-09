

import openpyxl
import sqlite3

global data_base_name


def phonebook_creator():
    data_base_name_input = input("Введіть назву для нової бази даних або вкажіть існуючу: ")
    global data_base_name
    data_base_name = "%s%s%s" % ('data/', data_base_name_input, '.db')
    conn = sqlite3.connect(data_base_name)
    phonebook_db = conn.cursor()
    with conn:
        try:
            phonebook_db.execute(""" CREATE TABLE phonebook_db (
                    id INTEGER PRIMARY KEY,
                    phone INTEGER,
                    name TEXT,
                    city TEXT,
                    address TEXT,
                    district TEXT
                )""")
            print('Базу', data_base_name, 'створено.')
        except sqlite3.OperationalError:
            print("Підключена база:", data_base_name)


def create_single_record():
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        input_attempt = 1
        print("Введіть '0' щоб повернутись у головне меню.")
        while input_attempt != 0:
            try:
                input_attempt -= 1
                user_input = input(
                    "Введіть дані через кому(5 значень без пропусків) ',' (Телефон,Ім'я,Місто,Адреса,Район): ")
                contact_input = list(user_input.split(','))
                if user_input == '0':
                    input_attempt = 0
                else:
                    if not contact_input[0].isnumeric() or len(contact_input[0]) != 12:
                        while not contact_input[0].isnumeric() or len(contact_input[0]) != 12:
                            if not contact_input[0].isnumeric():
                                contact_input[0] = input(
                                    "Номер телефона може бути тільки числовим, ще раз тільки номер: ")
                            elif len(contact_input[0]) != 12:
                                contact_input[0] = input(
                                    "Номер телефона має 12 символів починаючи з '38', ще раз тільки номер: ")
                    with conn:
                        phonebook_db.execute("""INSERT INTO phonebook_db (
                                                phone,
                                                name,
                                                city,
                                                address,
                                                district) 
                                                VALUES (?,?,?,?,?)""", contact_input)
                        print("Ви ввели: ", contact_input, "\nЗапис додано.")
            except sqlite3.ProgrammingError:
                input_attempt += 1
                print("Помилка вводу. Потрібно 5 значень через коми.")
    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)


def data_selector():
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        print("Доступні поля:\n"
              "[1] 'ID'\n"
              "[2] 'Телефон' або 'Phone'\n"
              "[3] 'Ім'я' або 'Name'\n"
              "[4] 'Місто' або 'City'\n"
              "[5] 'Адреса' або 'Address'\n"
              "[6] 'Район' або 'District'")
        user_attempt = 1
        search_column = ""
        counter = 0
        while user_attempt > 0:
            print("\nВведіть '0' для виходу в головне меню.")
            search_column_input = input("Введіть номер поля для пошуку: ")
            if int(search_column_input) == 0:
                user_attempt = 0
            elif int(search_column_input) > 6:
                print("Такого поля нема.")
            else:
                search_word = input("Введіть пошуковий термін: ")
                if int(search_column_input) == 1:
                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE id = {}".format(search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("№"),
                              '{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<27}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(counter),
                                  '{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<27}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0
                elif int(search_column_input) in range(2, 7):
                    if int(search_column_input) == 2:
                        search_column = 'phone'
                    elif int(search_column_input) == 3:
                        search_column = 'name'
                    elif int(search_column_input) == 4:
                        search_column = 'city'
                    elif int(search_column_input) == 5:
                        search_column = 'address'
                    elif int(search_column_input) == 6:
                        search_column = 'district'
                        print(search_column)

                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE {} LIKE '%{}%'"
                            .format(search_column, search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<30}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<30}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0
    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)


def edit_data_field():
    user_attempt = 1
    search_column = ""
    edit_attempt = 0
    edit_column = ""
    search_word = ""
    counter = 0
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        print("Доступні поля:\n"
              "[1] 'ID'\n"
              "[2] 'Телефон' або 'Phone'\n"
              "[3] 'Ім'я' або 'Name'\n"
              "[4] 'Місто' або 'City'\n"
              "[5] 'Адреса' або 'Address'\n"
              "[6] 'Район' або 'District'")
        while user_attempt == 1:
            print("\nВведіть '0' для виходу в головне меню.")
            search_column_input = input("Введіть номер поля для пошуку: ")
            if not search_column_input.isnumeric():
                print("Вибір поля може бути тільки числовим. Пункти меню від 1 до 6.")
                edit_attempt = -1
            elif int(search_column_input) == 0:
                user_attempt -= 1
            elif int(search_column_input) > 6:
                print("Такого поля нема.")
                edit_attempt = -1
            else:
                search_word = input("Введіть пошуковий термін: ")
                if int(search_column_input) == 1:
                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE id = {}".format(search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                        edit_attempt = -1
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<27}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<27}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0

                elif int(search_column_input) in range(2, 7):
                    if int(search_column_input) == 2:
                        search_column = 'phone'
                    elif int(search_column_input) == 3:
                        search_column = 'name'
                    elif int(search_column_input) == 4:
                        search_column = 'city'
                    elif int(search_column_input) == 5:
                        search_column = 'address'
                    elif int(search_column_input) == 6:
                        search_column = 'district'

                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE {} LIKE '%{}%'"
                            .format(search_column, search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                        edit_attempt -= 1
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<27}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<27}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0
                else:
                    pass

            edit_attempt += 1

            while edit_attempt == 1:
                print("\nДоступні поля:\n"
                      "[2] 'Телефон' або 'Phone'\n"
                      "[3] 'Ім'я' або 'Name'\n"
                      "[4] 'Місто' або 'City'\n"
                      "[5] 'Адреса' або 'Address'\n"
                      "[6] 'Район' або 'District'")
                print("\nВведіть '0' для виходу в головне меню.")
                edit_column_input = input("Введіть номер поля для редагування: ")
                if not edit_column_input.isnumeric():
                    print("Вибір поля може бути тільки числовим. Пункти меню від 1 до 6.")
                elif int(edit_column_input) == 0:
                    edit_attempt -= 1
                    user_attempt -= 1
                elif int(edit_column_input) == 1:
                    print("Поле ID недоступне для редагування.")
                elif int(edit_column_input) == 2:
                    print("\nВведіть '0' для повернення у головне меню.")
                    edit_column = 'phone'
                    updated_data = input("Введіть нове значення для поля: ")
                    edit_phone_attempt = 1
                    while not updated_data.isnumeric() or len(updated_data) != 12 and edit_phone_attempt == 1:
                        print ("\nВведіть '0' для повернення у головне меню.")
                        updated_data = input("Введіть нове значення для поля: ")
                        if not updated_data.isnumeric():
                            print("Номер телефона може бути тільки числовим, ще раз: ")
                        elif int(updated_data) == 0:
                            print("Вихід у гловне меню")
                            edit_phone_attempt -= 1
                            edit_attempt -= 1
                            user_attempt -= 1
                        elif len(updated_data) != 12:
                            print("Номер телефона має 12 символів починаючи з '38', ще раз: ")
                        else:
                            print("Невідома помилка вводу.")
                            edit_attempt -= 1
                            edit_phone_attempt -= 1
                            user_attempt -= 1

                elif int(edit_column_input) > 6:
                    print("Такого поля нема.")

                else:
                    print("\nВведіть '0' для виходу в головне меню.")
                    updated_data = input("Введіть нове значення для поля: ")
                    if int(updated_data) == 0:
                        print("Вихід у гловне меню")
                    else:
                        if int(edit_column_input) == 3:
                            edit_column = 'name'
                        elif int(edit_column_input) == 4:
                            edit_column = 'city'
                        elif int(edit_column_input) == 5:
                            edit_column = 'address'
                        elif int(edit_column_input) == 6:
                            edit_column = 'district'

                    with conn:
                        if int(search_column_input) == 1:
                            phonebook_db.execute("""
                            UPDATE phonebook_db SET {} = '{}' WHERE id = '{}'"""
                                                 .format(edit_column, updated_data, search_word))
                        else:
                            phonebook_db.execute("""
                            UPDATE phonebook_db SET {} = '{}' WHERE {} LIKE '%{}%'"""
                                                 .format(edit_column, updated_data, search_column, search_word))
                    print("Запис змінено.")
                    edit_attempt = 0
                    user_attempt = 0
    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)


def data_delete():
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        print("Доступні поля:\n"
              "[1] 'ID'\n"
              "[2] 'Телефон' або 'Phone'\n"
              "[3] 'Ім'я' або 'Name'\n"
              "[4] 'Місто' або 'City'\n"
              "[5] 'Адреса' або 'Address'\n"
              "[6] 'Район' або 'District'")
        user_attempt = 1
        search_column = ""
        counter = 0
        while user_attempt > 0:
            print("\nВведіть '0' для виходу в головне меню aбо 'format' для повного видалення бази")
            search_column_input = input("Введіть номер поля для пошуку: ")
            if not search_column_input.isnumeric():
                if search_column_input.upper() == 'FORMAT':
                    with conn:
                        phonebook_db.execute("SELECT * FROM phonebook_db")
                        results = phonebook_db.fetchall()
                        empty_base = []
                        if results == empty_base:
                            print("Даних по запиту не знайдено.")
                        else:
                            print("Телефонна книга:")
                            print('{:<5}'.format("ID"),
                                  '{:<15}'.format("Телефон"),
                                  '{:<20}'.format("Ім`я"),
                                  '{:<10}'.format("Місто"),
                                  '{:<27}'.format("Адреса"),
                                  '{:<30}'.format("Район"))
                            for item in results:
                                counter += 1
                                print('{:<5}'.format(item[0]),
                                      '{:<15}'.format(item[1]),
                                      '{:<20}'.format(item[2]),
                                      '{:<10}'.format(item[3]),
                                      '{:<27}'.format(item[4]),
                                      '{:<30}'.format(item[5]))
                            print("Всього записів:", counter)
                            confirmation = input("\nХочете видалити ці записи? (y/n): ")
                            if confirmation.upper() == 'Y':
                                user_attempt = 0
                                with conn:
                                    phonebook_db.execute("DELETE FROM phonebook_db")
                                print("Видалення виконано.")
                            else:
                                print('Видалення скасовано.')
                else:
                    print("Керування у меню можливо тільки числами.")
            elif int(search_column_input) == 0:
                user_attempt = 0
            elif int(search_column_input) > 6:
                print("Такого поля нема.")
            else:
                search_word = input("Введіть пошуковий термін: ")
                if int(search_column_input) == 1:
                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE id = {}".format(search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<27}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<27}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0
                        confirmation = input("\nХочете видалити ці записи? (y/n): ")
                        if confirmation.upper() == 'Y':
                            user_attempt = 0
                            with conn:
                                phonebook_db.execute(
                                    "DELETE FROM phonebook_db WHERE id = {}".format(search_word))
                            print("Видалення виконано.")
                        else:
                            print('Видалення скасовано.')
                elif int(search_column_input) in range(2, 7):
                    if int(search_column_input) == 2:
                        search_column = 'phone'
                    elif int(search_column_input) == 3:
                        search_column = 'name'
                    elif int(search_column_input) == 4:
                        search_column = 'city'
                    elif int(search_column_input) == 5:
                        search_column = 'address'
                    elif int(search_column_input) == 6:
                        search_column = 'district'
                        print(search_column)

                    with conn:
                        phonebook_db.execute(
                            "SELECT * FROM phonebook_db WHERE {} LIKE '%{}%'"
                            .format(search_column, search_word))
                    results = phonebook_db.fetchall()
                    empty_base = []
                    if results == empty_base:
                        print("Даних по запиту не знайдено.")
                    else:
                        print("Телефонна книга:")
                        print('{:<5}'.format("ID"),
                              '{:<15}'.format("Телефон"),
                              '{:<20}'.format("Ім`я"),
                              '{:<10}'.format("Місто"),
                              '{:<27}'.format("Адреса"),
                              '{:<30}'.format("Район"))
                        for item in results:
                            counter += 1
                            print('{:<5}'.format(item[0]),
                                  '{:<15}'.format(item[1]),
                                  '{:<20}'.format(item[2]),
                                  '{:<10}'.format(item[3]),
                                  '{:<27}'.format(item[4]),
                                  '{:<30}'.format(item[5]))
                        print("Всього записів:", counter)
                        counter = 0
                        confirmation = input("\nХочете видалити ці записи? (y/n): ")
                        if confirmation.upper() == 'Y':
                            user_attempt = 0
                            with conn:
                                phonebook_db.execute(
                                    "DELETE FROM phonebook_db WHERE {} LIKE '%{}%'".format(search_column, search_word)
                                )
                            print("Видалення виконано.")
                        else:
                            print('Видалення скасовано.')
                else:
                    print("Повторне введення.")
    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)


def sql_table_layout():
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        counter = 0
        with conn:
            phonebook_db.execute("SELECT * FROM phonebook_db")
            results = phonebook_db.fetchall()
            empty_base = []
            if results == empty_base:
                print("Даних по запиту не знайдено.")
            else:
                print("Телефонна книга:")
                print('{:<5}'.format("ID"),
                      '{:<15}'.format("Телефон"),
                      '{:<20}'.format("Ім`я"),
                      '{:<10}'.format("Місто"),
                      '{:<27}'.format("Адреса"),
                      '{:<30}'.format("Район"))
                for item in results:
                    counter += 1
                    print('{:<5}'.format(item[0]),
                          '{:<15}'.format(item[1]),
                          '{:<20}'.format(item[2]),
                          '{:<10}'.format(item[3]),
                          '{:<27}'.format(item[4]),
                          '{:<30}'.format(item[5]))
                print("Всього записів:", counter)

    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)


def xlsx_importer():
    try:
        conn = sqlite3.connect(data_base_name)
        phonebook_db = conn.cursor()
        print("Введіть назву .xlsx файлу, де є 5 колонок: ")
        print('{:<5}'.format("ID"),
              '{:<20}'.format("Телефон"),
              '{:<20}'.format("Імя"),
              '{:<10}'.format("Місто"),
              '{:<20}'.format("Адреса"),
              '{:<20}'.format("Район"))
        import_file = input("\nВведіть назву файлу для імпорту: ")
        import_file_name = "%s%s%s" % ('data/', import_file, '.xlsx')
        phonebook = openpyxl.load_workbook(import_file_name)
        table_1 = phonebook.active
        all_import = table_1.max_row
        print("Всього строк у файлі:", all_import - 1)
        contact_input = []
        line_index = 1
        import_counter = 0
        error_counter = 0
        while all_import - 1 > import_counter:
            with conn:
                for contact in table_1.iter_rows(min_row=2, min_col=1, max_col=5, max_row=all_import, values_only=True):
                    for cell in contact:
                        contact_input.append(cell)
                        phone = contact_input[0]
                    if not phone.isnumeric() and line_index in range(1, 5000, 5):
                        error_counter += 1
                        line_index += 5
                    else:
                        phonebook_db.execute("""INSERT INTO phonebook_db (
                                                phone,
                                                name,
                                                city,
                                                address,
                                                district) 
                                                VALUES (?,?,?,?,?)""", contact_input)
                    contact_input = []
                    import_counter += 1
                print(import_file_name, "імпортовано.")
                print("Вдалося імпортувати", all_import - error_counter - 1, "строк.")
                if all_import > all_import - error_counter:
                    print("Деякі строки мали букви у полі телефон, вони не були імпортовані.")

    except NameError:
        print("Бази даних не створено. Виберіть '0' у головному меню")
    except FileNotFoundError:
        print("Файлу не знайдено.")
    except sqlite3.OperationalError:
        print("Нема бази даних -", data_base_name)
